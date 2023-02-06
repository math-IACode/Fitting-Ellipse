import numpy as np
from scipy.optimize import minimize
from scipy.optimize import NonlinearConstraint, LinearConstraint
import matplotlib.pyplot as plt
import matplotlib.patches as pat
import matplotlib.transforms as transforms
import math
from openpyxl import load_workbook

vernal_equinox_angle=(25-21.64)/100*360
no_data_points=50
data_points=[]
column=36

#Importing data
data_file = 'Data_pointsnew.xlsx'
# Load the entire workbook.
wb = load_workbook(data_file)

# Load one worksheet.
ws = wb['Sheet1']
all_rows = list(ws.rows)
all_columns = list(ws.columns)

for row in range(6,154,3):
    cell=[[],[]]
    for i in range(3):
        v=ws.cell(row+i,column).value+ws.cell(row+i,column+1).value/60+ws.cell(row+i,column+2).value/(60*60)
        if v>180:
            cell[0].append(v-180)
        else:
            cell[0].append(v+180)
        cell[1].append(ws.cell(row+i,column+6).value+ws.cell(row+i,column+1+6).value/60+ws.cell(row+i,column+2+6).value/(60*60))
    data_points.append(cell)

# FUNCTIONS
def center_xy(val_A,val_B,val_C,val_D,val_E, num):
    if num==1:
        e_col = "cyan"
        s_col = "orange"
        m_col = "brown"
    else:
        e_col = "cyan"
        s_col = "orange"
        m_col = "brown"

    # Find center + change position of sun
    cent_x_val, cent_y_val = (2*val_C*val_D-val_E*val_B)/(val_B**2-4*val_A*val_C), (2*val_A*val_E-val_D*val_B)/(val_B**2-4*val_A*val_C)
    sun_x, sun_y = -cent_x_val, -cent_y_val
    
    # Draw earth orbit, and center points
    earth=plt.Circle((sun_x,sun_y),r,fill=False, color=e_col, label="Earth's orbit")
    ax.add_patch(earth)
    m = plt.scatter(0, 0, color=m_col, label="Location of Center of Mars Ellipse")
    #s = plt.scatter(sun_x,sun_y, color=s_col, label="Location of Sun")
    return cent_x_val, cent_y_val, sun_x, sun_y

def translate_eq(val_A,val_B,val_C,val_D,val_E,h,k):
    F = 1-(val_A*h**2+h*k*val_B+val_C*k**2+h*val_D+val_E*k)
    A=val_A/F
    B=val_B/F
    C=val_C/F
    D=(2*h*val_A+val_B*k+val_D)/F
    E=(h*val_B+2*k*val_C+val_E)/F
    return A,B,C,D,E

# General variables
fig, ax = plt.subplots()
plt.grid()
r=149.6
d=500

def position_of_mars(date,c,no):
    if no==33 or no==35:
        if no==33:
            p=2
        if no==35:
            p=0
        # Find coordinates of all points
        x1e, y1e = [0,r*math.cos(math.radians(date[0][1]+vernal_equinox_angle))], [0,r*math.sin(math.radians(date[0][1]+vernal_equinox_angle))] 
        x1m, y1m = [x1e[1],d*math.cos(math.radians(date[1][1]+vernal_equinox_angle))+x1e[1]], [y1e[1],d*math.sin(math.radians(date[1][1]+vernal_equinox_angle))+y1e[1]]

        x2e, y2e = [0,r*math.cos(math.radians(date[0][p]+vernal_equinox_angle))], [0,r*math.sin(math.radians(date[0][p]+vernal_equinox_angle))] 
        x2m, y2m = [x2e[1],d*math.cos(math.radians(date[1][p]+vernal_equinox_angle))+x2e[1]], [y2e[1],d*math.sin(math.radians(date[1][p]+vernal_equinox_angle))+y2e[1]]

        # Draw on screen
        #plt.plot(x1e,y1e,x1m,y1m, color=c,linewidth=0.5)
        #plt.plot(x2e,y2e,x2m,y2m,color=c,linewidth=0.5)

        m1m, m2m = (y1m[1]-y1m[0])/(x1m[1]-x1m[0]), (y2m[1]-y2m[0])/(x2m[1]-x2m[0])
        c1m, c2m = y1m[1]-m1m*x1m[1], y2m[1]-m2m*x2m[1]

        x_mid, y_mid = ((c1m-c2m)/(m2m-m1m)),((m1m*c2m-m2m*c1m)/(m1m-m2m))

        # Add the data points to list of data points
        X.append(x_mid)
        Y.append(y_mid)
        plt.plot(x_mid,y_mid, marker = 'x',markerfacecolor="black",markeredgecolor="black")

    else:
        # Find coordinates of all points
        x1e, y1e = [0,r*math.cos(math.radians(date[0][0]+vernal_equinox_angle))], [0,r*math.sin(math.radians(date[0][0]+vernal_equinox_angle))] 
        x1m, y1m = [x1e[1],d*math.cos(math.radians(date[1][0]+vernal_equinox_angle))+x1e[1]], [y1e[1],d*math.sin(math.radians(date[1][0]+vernal_equinox_angle))+y1e[1]]

        x2e, y2e = [0,r*math.cos(math.radians(date[0][1]+vernal_equinox_angle))], [0,r*math.sin(math.radians(date[0][1]+vernal_equinox_angle))] 
        x2m, y2m = [x2e[1],d*math.cos(math.radians(date[1][1]+vernal_equinox_angle))+x2e[1]], [y2e[1],d*math.sin(math.radians(date[1][1]+vernal_equinox_angle))+y2e[1]]

        x3e, y3e = [0,r*math.cos(math.radians(date[0][2]+vernal_equinox_angle))], [0,r*math.sin(math.radians(date[0][2]+vernal_equinox_angle))] 
        x3m, y3m = [x3e[1],d*math.cos(math.radians(date[1][2]+vernal_equinox_angle))+x3e[1]], [y3e[1],d*math.sin(math.radians(date[1][2]+vernal_equinox_angle))+y3e[1]]

        # Draw on screen
        #plt.plot(x1e,y1e,x1m,y1m, color=c,linewidth=0.5)
        #plt.plot(x2e,y2e,x2m,y2m,color=c,linewidth=0.5)
        #plt.plot(x3e,y3e,x3m,y3m,color=c,linewidth=0.5)

        # Find slope and y-intercept of each line
        m1m, m2m, m3m = (y1m[1]-y1m[0])/(x1m[1]-x1m[0]), (y2m[1]-y2m[0])/(x2m[1]-x2m[0]), (y3m[1]-y3m[0])/(x3m[1]-x3m[0])
        c1m, c2m, c3m = y1m[1]-m1m*x1m[1], y2m[1]-m2m*x2m[1], y3m[1]-m3m*x3m[1]

        # Find centroid of triangle (1:lines 1+2, 2:lines 2+3, 3:lines 1+3)
        xi1, xi2, xi3 = (c1m-c2m)/(m2m-m1m), (c2m-c3m)/(m3m-m2m), (c1m-c3m)/(m3m-m1m)
        yi1, yi2, yi3 = m1m*xi1+c1m, m2m*xi2+c2m, m3m*xi3+c3m

        # Find centroid of triangle
        x_mid = (xi1+xi2+xi3)/3
        y_mid = (yi1+yi2+yi3)/3

        # Add the data points to list of data points
        X.append(x_mid)
        Y.append(y_mid)

X=[]
Y=[]
# Print on screen - data points
for i in range(len(data_points)):
    position_of_mars(data_points[i], colours[i],i)

# .........................................................................................................................
# Initialise all matrices
X = np.array(X, ndmin=2)
Y = np.array(Y, ndmin=2)
X=X.reshape((no_data_points,1))
Y=Y.reshape((no_data_points,1))
plt.scatter(X,Y, marker='x',color='black')

# .........................................................................................................................
# Fitting number 1 
XY_coord=[X,Y]
XY_coord = np.array(XY_coord)
XY_coord=XY_coord.reshape((2,no_data_points))

# Formulate and solve the least squares problem ||Ax - b ||^2
A1 = np.hstack([X**2, Y**2, X, Y])
b1 = np.ones_like(X)
x1 = np.linalg.lstsq(A1, b1)[0].squeeze()
x1=x1.reshape((4,1))

# Calculate cost function
cost1=np.sum(np.square(np.subtract(np.matmul(A1,x1), b1)))

# Calculate center + translate ellipse
x_1_center_cal, y_1_center_cal, x_1_loc_sun, y_1_loc_sun = center_xy(x1[0],0,x1[1],x1[2],x1[3], 1)
x1[0],NNN, x1[1],x1[2],x1[3]=translate_eq(x1[0],0,x1[1],x1[2],x1[3],x_1_center_cal,y_1_center_cal)

# Plot the least squares ellipse
x_coord = np.linspace(-300,300,300)
y_coord = np.linspace(-300,300,300)
X_coord, Y_coord = np.meshgrid(x_coord, y_coord)
Z_coord = x1[0,0] * X_coord ** 2 + x1[1,0] * Y_coord**2 + x1[2,0] * X_coord + x1[3,0] * Y_coord
plt.contour(X_coord, Y_coord, Z_coord, levels=[1], colors=('r'), linewidths=2)
print('The ellipse is given by {0:.3}x^2+{1:.3}y^2+{2:.3}x+{3:.3}y = 1'.format(x1[0,0], x1[1,0],x1[2,0],x1[3,0]))

# .........................................................................................................................
# Formulate and solve the least squares problem ||Ax - b ||^2
A2 = np.hstack([X**2, X*Y, Y**2, X, Y])
b2 = np.ones_like(X)
x2 = np.linalg.lstsq(A2, b2)[0].squeeze()
x2=x2.reshape((5,1))

Z_coord = x2[0,0] * X_coord ** 2 + x2[1,0] * X_coord*Y_coord + x2[2,0] * Y_coord**2 + x2[3,0] * X_coord + x2[4,0] * Y_coord
#plt.contour(X_coord, Y_coord, Z_coord, levels=[1], colors=('b'), linewidths=2)

# Calculate cost function
cost2=np.sum(np.square(np.subtract(np.matmul(A2,x2), b2)))

# Calculate center + translate ellipse
x_2_center_cal, y_2_center_cal, x_2_loc_sun, y_2_loc_sun = center_xy(x2[0],x2[1],x2[2],x2[3], x2[4], 2)
x2[0],x2[1],x2[2],x2[3],x2[4]=translate_eq(x2[0],x2[1],x2[2],x2[3],x2[4],x_2_center_cal,y_2_center_cal)

# Plt unrotated curve
Z_coord = x2[0,0] * X_coord ** 2 + x2[1,0]* X_coord * Y_coord + x2[2,0] * Y_coord**2 + x2[3,0] * X_coord + x2[4,0] * Y_coord
plt.contour(X_coord, Y_coord, Z_coord, levels=[1], colors=('g'), linewidths=2)
print('The ellipse is given by {0:.3}x^2+{1:.3}xy+{2:.3}y^2+{3:.3}x+{4:.3}y = 1'.format(x2[0,0], x2[1,0],x2[2,0],x2[3,0],x2[4,0]))

# Find actual equaion without
angle_of_rotation=np.arctan(x2[1,0]/(x2[0,0]-x2[2,0]))/2
print(angle_of_rotation*180/np.pi)
A=x2[0,0]*np.cos(angle_of_rotation)**2+x2[1,0]*np.sin(angle_of_rotation)*np.cos(angle_of_rotation)+x2[2,0]*np.sin(angle_of_rotation)**2
B=(x2[2,0]-x2[0,0])*np.sin(2*angle_of_rotation)+x2[1,0]*np.cos(2*angle_of_rotation)
C=x2[0,0]*np.sin(angle_of_rotation)**2-x2[1,0]*np.sin(angle_of_rotation)*np.cos(angle_of_rotation)+x2[2,0]*np.cos(angle_of_rotation)**2
D=x2[3,0]*np.cos(angle_of_rotation)+x2[4,0]*np.sin(angle_of_rotation)
E=x2[4,0]*np.cos(angle_of_rotation)-x2[3,0]*np.sin(angle_of_rotation)
F=-1
print(A,B,C,D,E,F)
# Turn new coefficients into a matrix
x3=np.array([A,C,D,E])
x3=x3.reshape((4,1))

# .........................................................................................................................
# Conversion to other equation

# First ellipse
h1 = -x1[2,0]/(2*x1[0,0])
k1 = -x1[3,0]/(2*x1[1,0])
a1 = np.sqrt(1/x1[0,0]+x1[2,0]**2/(4*x1[0,0]**2)+x1[3,0]**2/(4*x1[0,0]*x1[1,0])) # semi-minor
b1 = np.sqrt(1/x1[1,0]+x1[2,0]**2/(4*x1[0,0]*x1[1,0])+x1[3,0]**2/(4*x1[1,0]**2)) # semi-major
c1 = np.sqrt(b1**2-a1**2) # foci points
e1 = c1/b1
x1_c1 = h1
y1_c1 = k1+c1
x2_c1 = h1
y2_c1 = k1-c1
plt.scatter([x1_c1, x2_c1],[y1_c1, y2_c1], color="r", label="foci of red ellipse")
distance_sun_foc1 = np.sqrt((y2_c1-y_1_loc_sun)**2+(x2_c1-x_1_loc_sun)**2)

# Second ellipse
h2 = -x3[2,0]/(2*x3[0,0])
k2 = -x3[3,0]/(2*x3[1,0])
h2,k2 = h2*np.cos(angle_of_rotation)-k2*np.sin(angle_of_rotation), h2*np.sin(angle_of_rotation)+k2*np.cos(angle_of_rotation)
a2 = np.sqrt(1/x3[0,0]+x3[2,0]**2/(4*x3[0,0]**2)+x3[3,0]**2/(4*x3[0,0]*x3[1,0])) # semi-minor
b2 = np.sqrt(1/x3[1,0]+x3[2,0]**2/(4*x3[0,0]*x3[1,0])+x3[3,0]**2/(4*x3[1,0]**2)) # semi-major
c2 = np.sqrt(b2**2-a2**2) # foci points
e2 = c2/b2
x1_c2 = c2*np.cos(angle_of_rotation)
y1_c2 = c2*np.sin(angle_of_rotation)
x2_c2 = -c2*np.cos(angle_of_rotation)
y2_c2 = -c2*np.sin(angle_of_rotation)
plt.scatter([x1_c2, x2_c2],[y1_c2, y2_c2], color="g",label="foci of blue ellipse")
distance_sun_foc2 = np.sqrt((y1_c2-y_2_loc_sun)**2+(x1_c2-x_2_loc_sun)**2)

print(e1,e2)
print(h1,k1,a1,b1)
print(x_1_loc_sun,x_2_loc_sun,y_1_loc_sun,y_2_loc_sun)
print("foci points:",str([x1_c1,y1_c1,x2_c1,y2_c1]))
print(h2,k2,a2,b2)
print("foci points:",str([x1_c2,y1_c2,x2_c2,y2_c2]))
print("distance from foci to sun", str([distance_sun_foc1, distance_sun_foc2]))

# .........................................................................................................................
# Method 3

# Position of center and sun location
x_center, y_center = 0,0
x_sun, y_sun = (x_1_loc_sun+x_2_loc_sun)/2, (y_1_loc_sun+y_2_loc_sun)/2

# Find c
dist_center_to_sun = np.sqrt(x_sun**2 + y_sun**2)


# Find tilt angle 
angle_tilt_sun = np.arctan(y_sun/x_sun)

# Move points by c and then rotate by tilt angle anticlockwise
## Shift points 
shift_x_direction = np.repeat(x_sun, no_data_points)
shift_y_direction = np.repeat(y_sun, no_data_points)
shift_x_direction = shift_x_direction.reshape((no_data_points,1))
shift_y_direction = shift_y_direction.reshape((no_data_points,1))
X_new = np.add(X, shift_x_direction)
Y_new = np.add(Y, shift_y_direction)

## Rotate points
XY_coord=[X_new,Y_new]
XY_coord = np.array(XY_coord)
XY_coord=XY_coord.reshape((2,no_data_points))
Xrotation_matrix=np.array([np.cos(angle_tilt_sun), np.sin(angle_tilt_sun)])
Xrotation_matrix=Xrotation_matrix.reshape((1,2))
Yrotation_matrix=np.array([-np.sin(angle_tilt_sun), np.cos(angle_tilt_sun)])
Yrotation_matrix=Yrotation_matrix.reshape((1,2))
X_new=np.matmul(Xrotation_matrix, XY_coord)
Y_new=np.matmul(Yrotation_matrix, XY_coord)
X_new=X_new.reshape((no_data_points,1))
Y_new=Y_new.reshape((no_data_points,1))

## Rotate sun 
x_sun, y_sun = x_sun*np.cos(angle_tilt_sun)+y_sun*np.sin(angle_tilt_sun), -x_sun*np.sin(angle_tilt_sun)+y_sun*np.cos(angle_tilt_sun)

#plt.scatter(X_new,Y_new,marker="x")
plt.scatter(x_sun,y_sun, color="orange")

# Formulate and solve the least squares problem ||Ax - b ||^2
A2 = np.hstack([(-(dist_center_to_sun**2-Y_new**2-X_new**2)+np.sqrt((dist_center_to_sun**2-X_new**2-Y_new**2)**2-4*(dist_center_to_sun**2)*Y_new**2))/2])
b2 = np.repeat(1,no_data_points)
x2 = np.linalg.lstsq(A2, b2)[0].squeeze()
x2=x2.reshape((1,1))
print(dist_center_to_sun)
b_squared=1/x2
a_squared=dist_center_to_sun**2+b_squared
Z_coord = 1/(a_squared) * X_coord ** 2 + 1/b_squared * Y_coord**2
plt.contour(X_coord, Y_coord, Z_coord, levels=[1], colors=('b'), linewidths=2)
print(np.sqrt(a_squared), np.sqrt(b_squared))
cost3=np.sum(np.square(np.subtract(np.matmul(A2,x2), b2)))

# .........................................................................................................................
# Print/Plot everything
print('The ellipse is given by {0:.3}x^2+{1:.3}y^2+{2:.3}x+{3:.3}y = 1'.format(x1[0,0], x1[1,0],x1[2,0],x1[3,0]))
print('The ellipse is given by {0:.3}x^2+{1:.3}y^2+{2:.3}x+{3:.3}y = 1'.format(x3[0,0], x3[1,0],x3[2,0],x3[3,0]))
print("costs:", str([cost1, cost2,cost3]))

plt.title("Fitting Ellipses")
plt.xlabel("Distance (x) in million km")
plt.ylabel("Distance (y) in million km")
ax.set_aspect('equal', adjustable='box')
plt.show()
